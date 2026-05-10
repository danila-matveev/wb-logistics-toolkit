# tests/localization/output/test_writer.py
from unittest.mock import patch

from openpyxl import load_workbook

from localization.output.writer import ExcelWriter, SheetsWriter, make_writer


def test_excel_writer_creates_file_with_multiple_sheets(tmp_path):
    out = tmp_path / "test.xlsx"
    w = ExcelWriter(str(out))
    w.write_sheet("Анализ", [["Артикул", "Цена"], ["ABC", 100], ["DEF", 200]])
    w.write_sheet("Сводка", [["Метрика", "Значение"], ["ИЛ %", 75.5]])
    path = w.finalize()

    assert path == str(out)
    assert out.exists()

    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Анализ", "Сводка"}
    assert wb["Анализ"].cell(1, 1).value == "Артикул"
    assert wb["Анализ"].cell(2, 2).value == 100
    assert wb["Сводка"].cell(2, 2).value == 75.5


def test_excel_writer_overwrites_existing_file(tmp_path):
    out = tmp_path / "exists.xlsx"
    out.write_bytes(b"junk")
    w = ExcelWriter(str(out))
    w.write_sheet("OnlySheet", [["A"], ["B"]])
    w.finalize()
    wb = load_workbook(out)
    assert wb.sheetnames == ["OnlySheet"]


def test_excel_writer_handles_empty_data(tmp_path):
    out = tmp_path / "empty.xlsx"
    w = ExcelWriter(str(out))
    w.write_sheet("Empty", [])
    w.finalize()
    wb = load_workbook(out)
    assert wb["Empty"].max_row == 1   # openpyxl creates 1 phantom row


def test_excel_writer_creates_parent_dir(tmp_path):
    out = tmp_path / "nested" / "deep" / "out.xlsx"
    w = ExcelWriter(str(out))
    w.write_sheet("S", [["x"]])
    w.finalize()
    assert out.exists()


def test_make_writer_picks_excel_when_no_sheet_id(tmp_path):
    w = make_writer(
        sheet_id="",
        excel_path=str(tmp_path / "out.xlsx"),
        force_excel=False,
    )
    assert isinstance(w, ExcelWriter)


def test_make_writer_picks_excel_when_placeholder_sheet_id(tmp_path):
    w = make_writer(
        sheet_id="YOUR_SHEET_ID_HERE",
        excel_path=str(tmp_path / "out.xlsx"),
        force_excel=False,
    )
    assert isinstance(w, ExcelWriter)


def test_make_writer_picks_excel_when_creds_missing(tmp_path, monkeypatch):
    monkeypatch.setenv("GOOGLE_CREDENTIALS_PATH", str(tmp_path / "missing.json"))
    w = make_writer(
        sheet_id="1AbCxyz",
        excel_path=str(tmp_path / "out.xlsx"),
        force_excel=False,
    )
    assert isinstance(w, ExcelWriter)


def test_make_writer_force_excel_overrides_valid_config(tmp_path, monkeypatch):
    creds = tmp_path / "creds.json"
    creds.write_text("{}")
    monkeypatch.setenv("GOOGLE_CREDENTIALS_PATH", str(creds))
    w = make_writer(
        sheet_id="1AbCxyz",
        excel_path=str(tmp_path / "out.xlsx"),
        force_excel=True,
    )
    assert isinstance(w, ExcelWriter)


def test_make_writer_picks_sheets_when_all_conditions_met(tmp_path, monkeypatch):
    creds = tmp_path / "creds.json"
    creds.write_text("{}")
    monkeypatch.setenv("GOOGLE_CREDENTIALS_PATH", str(creds))
    fake_spreadsheet = object()
    with patch(
        "localization.output.writer._open_spreadsheet",
        return_value=fake_spreadsheet,
    ) as mock_open:
        w = make_writer(
            sheet_id="1AbCxyz",
            excel_path=str(tmp_path / "out.xlsx"),
            force_excel=False,
        )
    assert isinstance(w, SheetsWriter)
    mock_open.assert_called_once_with("1AbCxyz")

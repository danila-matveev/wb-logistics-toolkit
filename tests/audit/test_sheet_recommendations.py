from datetime import date
import openpyxl
from audit.output.sheet_recommendations import write_recommendations
from audit.models.report_row import ReportRow
from audit.calculators.logistics_overpayment import OverpaymentResult


def _make_row(nm_id: int, office: str, delivery_rub: float, order_dt: date) -> ReportRow:
    return ReportRow.from_api({
        "nm_id": nm_id,
        "office_name": office,
        "supplier_oper_name": "Логистика",
        "bonus_type_name": "К клиенту при продаже",
        "delivery_rub": delivery_rub,
        "order_dt": order_dt.isoformat(),
        "realizationreport_id": 1,
    })


def test_write_recommendations_total_row():
    wb = openpyxl.Workbook()
    ws = wb.active
    rows = [
        _make_row(1, "Коледино", 100.0, date(2026, 1, 10)),
        _make_row(2, "Тула", 200.0, date(2026, 2, 5)),
    ]
    results = [
        OverpaymentResult(calculated_cost=60.0, overpayment=40.0),
        OverpaymentResult(calculated_cost=100.0, overpayment=100.0),
    ]
    write_recommendations(ws, rows, results, date(2026, 1, 1), date(2026, 3, 31))
    # Check total overpayment cell exists and is 140
    values = [[ws.cell(r, c).value for c in range(1, 4)] for r in range(1, 20)]
    flat = [v for row in values for v in row if v is not None]
    assert 140.0 in flat or "140.0" in [str(v) for v in flat]


def test_write_recommendations_header_row():
    wb = openpyxl.Workbook()
    ws = wb.active
    write_recommendations(ws, [], [], date(2026, 1, 1), date(2026, 3, 31))
    assert ws.cell(1, 1).value is not None  # has a header

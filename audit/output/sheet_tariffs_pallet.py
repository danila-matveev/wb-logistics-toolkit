"""Sheet 11: 'Тариф монопалета' — pallet tariff data."""
from __future__ import annotations
from openpyxl.worksheet.worksheet import Worksheet


def write_tariffs_pallet(ws: Worksheet, pallet_data: list[dict]) -> None:
    """Write raw pallet tariff data. `pallet_data` is the warehouseList already
    unpacked by shared.wb_api.tariffs.fetch_pallet_tariffs (a flat list of dicts)."""
    if not pallet_data:
        ws.cell(1, 1, "Нет данных по тарифам монопалета")
        return
    headers = list(pallet_data[0].keys())
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    for i, wh in enumerate(pallet_data, 2):
        for col, h in enumerate(headers, 1):
            ws.cell(i, col, wh.get(h, ""))

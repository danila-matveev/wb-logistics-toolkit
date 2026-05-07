"""Sheet 12: 'Рекомендации' — overpayment summary for claims preparation."""
from __future__ import annotations
from collections import defaultdict
from datetime import date

from openpyxl.worksheet.worksheet import Worksheet

from audit.models.report_row import ReportRow
from audit.calculators.logistics_overpayment import OverpaymentResult


def write_recommendations(
    ws: Worksheet,
    logistics_rows: list[ReportRow],
    overpayment_results: list[OverpaymentResult | None],
    date_from: date,
    date_to: date,
) -> None:
    """Write summary for claims/lawsuit preparation.

    Sections:
    1. Overall totals
    2. Top-10 articles by overpayment sum
    3. Top-5 warehouses by overpayment sum
    4. Monthly overpayment trend
    """
    # Aggregate
    total_charged = 0.0
    total_overpay = 0.0
    by_article: dict[int, float] = defaultdict(float)
    by_warehouse: dict[str, float] = defaultdict(float)
    by_month: dict[str, float] = defaultdict(float)

    for row, res in zip(logistics_rows, overpayment_results):
        if res is None or res.overpayment < 0:
            continue
        total_charged += row.delivery_rub
        total_overpay += res.overpayment
        by_article[row.nm_id] += res.overpayment
        by_warehouse[row.office_name] += res.overpayment
        if row.order_dt:
            month_key = row.order_dt.strftime("%Y-%m")
            by_month[month_key] += res.overpayment

    excel_row = 1

    # Section 1: Overall totals
    ws.cell(excel_row, 1, "ОБЩИЙ ИТОГ")
    excel_row += 1
    ws.cell(excel_row, 1, "Период")
    ws.cell(excel_row, 2, f"{date_from.isoformat()} — {date_to.isoformat()}")
    excel_row += 1
    ws.cell(excel_row, 1, "WB удержал за логистику (₽)")
    ws.cell(excel_row, 2, round(total_charged, 2))
    excel_row += 1
    ws.cell(excel_row, 1, "Переплата по нашему расчёту (₽)")
    ws.cell(excel_row, 2, round(total_overpay, 2))
    excel_row += 1
    pct = total_overpay / total_charged * 100 if total_charged else 0
    ws.cell(excel_row, 1, "Доля переплаты (%)")
    ws.cell(excel_row, 2, round(pct, 1))
    excel_row += 2

    # Section 2: Top-10 articles
    ws.cell(excel_row, 1, "ТОП-10 АРТИКУЛОВ ПО СУММЕ ПЕРЕПЛАТЫ")
    excel_row += 1
    ws.cell(excel_row, 1, "Код номенклатуры")
    ws.cell(excel_row, 2, "Переплата (₽)")
    excel_row += 1
    for nm_id, overpay in sorted(by_article.items(), key=lambda x: x[1], reverse=True)[:10]:
        ws.cell(excel_row, 1, nm_id)
        ws.cell(excel_row, 2, round(overpay, 2))
        excel_row += 1
    excel_row += 1

    # Section 3: Top-5 warehouses
    ws.cell(excel_row, 1, "ТОП-5 СКЛАДОВ ПО СУММЕ ПЕРЕПЛАТЫ")
    excel_row += 1
    ws.cell(excel_row, 1, "Склад")
    ws.cell(excel_row, 2, "Переплата (₽)")
    excel_row += 1
    for wh, overpay in sorted(by_warehouse.items(), key=lambda x: x[1], reverse=True)[:5]:
        ws.cell(excel_row, 1, wh)
        ws.cell(excel_row, 2, round(overpay, 2))
        excel_row += 1
    excel_row += 1

    # Section 4: Monthly trend
    ws.cell(excel_row, 1, "ДИНАМИКА ПЕРЕПЛАТ ПО МЕСЯЦАМ")
    excel_row += 1
    ws.cell(excel_row, 1, "Месяц")
    ws.cell(excel_row, 2, "Переплата (₽)")
    excel_row += 1
    for month_key in sorted(by_month):
        ws.cell(excel_row, 1, month_key)
        ws.cell(excel_row, 2, round(by_month[month_key], 2))
        excel_row += 1
